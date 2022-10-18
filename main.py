import openpyxl
import pandas as pd
from openpyxl import load_workbook
import re

# load up the excel docs in pandas dataframes
contract_summary_df = data = pd.read_excel("clm-contract-summary-report_final.xlsx")
line_of_accounting_df = data = pd.read_excel("line-of-accouting-training-boc.xlsx")

def extract_categories():
    # create empty dicts to store extracted data
    expenditure_dict = {}
    account_dict = {}

    # iterate over rows in the contract summary dataframe
    for index, row in contract_summary_df.iterrows():
        # access expenditure description value in the row
        expenditure_description = str(row["Expenditure Type"])

        # check if the row has a description
        if len(expenditure_description) > 0:

            # search category numbers from contract summary df with regex and lookup in accounts df
            category_number = re.search("^[0-9]{3}", expenditure_description)
            if category_number is not None:
                category_number = int(category_number.group(0))

                # for float numbers
                if category_number % 10 != 0:
                    category_number = category_number / 10

                    # lookup
                    category_name = line_of_accounting_df.loc[line_of_accounting_df["Value"] == category_number]["Description"].values[0]
                    
                else:
                     # for integer numbers
                    category_number = int(category_number/10)

                    # lookup
                    category_name = line_of_accounting_df.loc[line_of_accounting_df["Value"] == category_number]["Description"].values[0]
                    
            else:

                # set to empty if there is no value
                category_name = ""

            # Set sub category names
            # similar operation for sub category names
            sub_category_number = re.search("^[0-9]+[A-Z]*", expenditure_description)
            if sub_category_number is not None:
                sub_category_number = sub_category_number.group(0)
                
                try:
                    sub_category_number = int(sub_category_number)
                    
                except:
                    pass

                try:
                    sub_category_name = line_of_accounting_df.loc[line_of_accounting_df["Value"] == sub_category_number]["Description"].values[0]

                except:
                    sub_category_name = ""
                            
            else:
                sub_category_name = ""

        else:
            category_name = ""
            sub_category_name = ""
            
        expenditure_dict[index] = [category_name, sub_category_name]

        # funding account
        try:
            # Split into individual values
            account_dict[index] = row["Funding Account"].split("|")

        except:
            account_dict[index] = ""

        print(f"Reading row {index} of {contract_summary_df.shape[0]}")

    # open excel sheet and set active for editing
    contracts = openpyxl.load_workbook("clm-contract-summary-report_final.xlsx")
    sheet = contracts.active

    no_of_rows = len(expenditure_dict)

    # set all values as required
    for indx, val in expenditure_dict.items():
        try:
            sheet[f"O{indx+2}"].value = val[0]
            sheet[f"P{indx+2}"].value = val[1]

        except:
            sheet[f"O{indx+2}"].value = val[0]
            sheet[f"P{indx+2}"].value = val[1]

        for i, v in enumerate(account_dict[indx]):
            if indx == 0:
                sheet.cell(row=indx+1, column=i+25).value = f"Column {i+1}"
            sheet.cell(row=indx+2, column=i+25).value= v

        print(f"writing row {indx} of {no_of_rows}")
    
    # save file
    contracts.save("clm-contract-summary-report_final.xlsx")

extract_categories()